#!/usr/bin/env python3
import sqlite3
import argparse
import os
import sys
import re
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill

version = "v2.1c, 20230702"

today = date.today()
dir = today.strftime("%Y%m%d")

parser = argparse.ArgumentParser(description='creates an index.html page from the sqlite database websites.db')
parser.add_argument('-d','--debug', action='store_true', help='print debug messages to stderr')
parser.add_argument('-p', '--path', type=str, help=f'The directory path, if not given will assume the directory with todays date: {dir}')
parser.add_argument('-m', '--mail', type=str, help='e-mail address to send the Excel file to')
parser.add_argument('-v','--version', action='store_true', help='show version info and exit')

args = parser.parse_args()
debug = args.debug

if args.version:
    sys.exit(f"version: {version}")

# If the path argument is provided, use it as the directory path
if args.path:
    directory_path = args.path
else:
    # If the path argument is not provided, set the directory name to today's date
    directory_path = dir

debug and print("debug output activated")
debug and print(f"Will read from, and store into directory: {directory_path}")

# Check if the directory exists
if not os.path.exists(directory_path):
    print(f"The directory {directory_path} does not exist.")
    exit()

# Connect to the SQLite database in the directory
try:
    database = os.path.join(directory_path, 'websites.db')
    conn = sqlite3.connect(database)
    debug and print(f"Connected to database {database}")
    c = conn.cursor()
except sqlite3.Error as e:
    sys.exit(f"Error connecting to database {database}: {e}")

# headers as a list, in the order they should appear in the excel sheet
table_headers = ['website', 
                 'https',
                 'grade', 
                 'grade check',
                 'HTTPS redirect',
                 'certificate validity',
                 'HSTS',
                 'headers', 
                 'security.txt', 
                 'robots.txt', 
                 'version', 
                 'error', 
                 'remnants',
                 'debug',
                 'check date']

# Query the table structure from the meta table
c.execute("SELECT structure FROM meta")
result = c.fetchone()

if result:
    table_structure = result[0]
    debug and print(f"Table structure: {table_structure}")

    # Extract column names from the table structure
    column_pattern = re.compile(r'(\w+)\s+[\w\(\)]+(,)?')
    columns = [match.group(1) for match in column_pattern.finditer(table_structure) if match.group(1) != "IF"]

    # Build the SQL query
    sql_query = "SELECT {} FROM website_checks".format(", ".join(columns))
    debug and print(f"sql_query = {sql_query}")
else:
    sys.exit(f"unable to read database structure from {database}")

cs = conn.cursor()
try:
    cs.execute(sql_query)
    rows = cs.fetchall()
except sqlite3.Error as error:
    sys.exit(f"Failed to fetch data from {database}", error)


# Create a new workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Define cell styles for headers, OK values, and NotOK values
header_style = Font(bold=True)
value_ok_style = PatternFill(start_color='00FF89', end_color='00FF89', fill_type='solid')
value_notok_style = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

for col, header in enumerate(table_headers):
    worksheet.cell(row=1, column=col+1, value=header).font = header_style
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = 15

# Write the data rows with appropriate styles
for row, data in enumerate(rows):
    for col, value in enumerate(data):
        debug and print(f"row = {row}, col = {col}, value = {value}")
           
        cell = worksheet.cell(row=row+2, column=col+1, value=value)

        # col 0 = website name
        # col 2 = grade
        # col 5 = cert. validity (days)
        # col 6 = hsts duration
        # col 14 = check_date
        if col in [1,3,4,7,8,9,10,11,12,13]:
            if value == 1:
                cell.fill = value_ok_style
                cell.value = 'OK'
            else:
                cell.fill = value_notok_style
                cell.value = 'NotOK'

# Save the excel workbook
excelfile = os.path.join(directory_path, 'website_checks.xlsx')
try:
    workbook.save(excelfile)
    print(f"saving to: {excelfile}")
except OSError as e:
    sys.exit(f"Error saving excel file: {e}")

# If the -m switch was specified, e-mail the Excel file to the specified address
# N.B. this has to be modified to your specific situation before it will actually send an e-mail
if args.mail:
    msg = MIMEMultipart()
    msg['From'] = 'sender@example.com'
    msg['To'] = args.mail
    msg['Subject'] = 'Website Checks Report'
    with open(excelfile, 'rb') as f:
        attach = MIMEApplication(f.read(), _subtype = 'xls')
        attach.add_header('Content-Disposition', 'attachment', filename='website_checks.xlsx')
        msg.attach(attach)
    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    smtp.starttls()
    smtp.login('sender@example.com', 'password')        # N.B. be very careful with putting passwords in scripts
    smtp.sendmail('sender@example.com', args.mail, msg.as_string())
    smtp.quit()
